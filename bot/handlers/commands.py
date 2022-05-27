import random

from aiogram import types, Dispatcher
from openpyxl import load_workbook
from sqlalchemy import select, func

from bot.db.models import PlayerScore, User, Role, Lang, Word, Tag
from bot.db_tools import get_or_create
from bot.keyboards import generate_balls, generate_question


async def cmd_install(message: types.Message):
    admin_id = message.bot.get("admin_id")
    current_id = message.from_user.id

    if current_id not in admin_id:
        return
    db_session = message.bot.get("db")
    admin = await get_or_create(db_session, Role, name='Admin', code='admin')
    premium = await get_or_create(db_session, Role, name='Premium', code='premium')
    subscriber = await get_or_create(db_session, Role, name='Subscriber', code='subscriber')
    user = await get_or_create(db_session, User, id=message.from_user.id, role_id=admin.id)
    await get_or_create(db_session, Lang, name='English', code='en')
    ru_lang = await get_or_create(db_session, Lang, name='Русский', code='ru')
    tag = await get_or_create(db_session, Tag, label='base')
    await get_or_create(
        db_session,
        Word,
        translate_id=ru_lang.id,
        tag_id=tag.id,
        original='turn',
        transcription='[tɜːn]',
        translate='поворачивать')
    await get_or_create(
        db_session,
        Word,
        translate_id=ru_lang.id,
        tag_id=tag.id,
        original='incorporate',
        transcription='[ɪnˈkɔːpəreɪt]',
        translate='включать (в себя что-то), объединяться')
    await message.answer("ok!")


async def cmd_start(message: types.Message):
    await message.answer("Hello!")
    db_session = message.bot.get("db")
    admin_id = message.bot.get("admin_id")
    current_id = message.from_user.id
    if current_id not in admin_id:
        role = await get_or_create(db_session, Role, name='Subscriber', code='subscriber')
    else:
        role = await get_or_create(db_session, Role, name='Admin', code='admin')
    user = await get_or_create(db_session, User, id=message.from_user.id, role=role)
    if user:
        await message.answer(f"you id - {user.id}, you role - {role.name}!")


async def cmd_play(message: types.Message):
    db_session = message.bot.get("db")

    async with db_session() as session:
        await session.merge(PlayerScore(user_id=message.from_user.id, score=0))
        await session.commit()

    await message.answer("Your score: 0", reply_markup=generate_balls())


async def cmd_read_xlsx1(message: types.Message):
    db_session = message.bot.get("db")

    workbook = load_workbook('./datastore/data.xlsx')
    worksheets = workbook.get_sheet_names()

    for worksheet_name in worksheets:
        worksheet = workbook.get_sheet_by_name(worksheet_name)

        tag = await get_or_create(db_session, Tag, label='base')
        for row in worksheet.iter_rows():
            ol, ts, tt = row
            ru_lang = await get_or_create(db_session, Lang, name='Русский', code='ru')
            if ol.value is None:
                continue

            if ts.value is None and tt.value is None:
                tag = await get_or_create(db_session, Tag, label=f"{ol.value} [{worksheet_name}]")

                continue
    await message.answer("imported")


async def cmd_read_xlsx2(message: types.Message):
    db_session = message.bot.get("db")

    workbook = load_workbook('./datastore/data.xlsx')
    worksheets = workbook.get_sheet_names()

    for worksheet_name in worksheets:
        worksheet = workbook.get_sheet_by_name(worksheet_name)

        tag = await get_or_create(db_session, Tag, label='base')
        for row in worksheet.iter_rows():
            ol, ts, tt = row
            ru_lang = await get_or_create(db_session, Lang, name='Русский', code='ru')
            if ol.value is None:
                continue

            if ts.value is None and tt.value is None:
                tag = await get_or_create(db_session, Tag, label=f"{ol.value} [{worksheet_name}]")

                continue

            await get_or_create(
                db_session,
                Word,
                original=ol.value,
                translate_id=ru_lang.id,
                transcription=ts.value,
                translate=tt.value,
                tag_id=tag.id)
    # async with db_session() as session:
    #     await session.merge(PlayerScore(user_id=message.from_user.id, score=0))
    #     await session.commit()

    # await message.answer("Your score: 0", reply_markup=generate_balls())
    await message.answer("imported")


async def get_card(message: types.Message):
    db_session = message.bot.get("db")
    sql = select(Word).order_by(func.random()).limit(4)
    async with db_session() as session:
        words_request = await session.execute(sql)
        words = words_request.scalars().all()
    true_answer = words[random.randint(0, 3)]
    await message.answer(
        f"{true_answer.original}  {true_answer.transcription}",
        reply_markup=generate_question(words, true_answer))


async def cmd_top(message: types.Message):
    db_session = message.bot.get("db")
    sql = select(PlayerScore).order_by(PlayerScore.score.desc()).limit(5)
    text_template = "Top 5 players:\n\n{scores}"
    async with db_session() as session:
        top_players_request = await session.execute(sql)
        players = top_players_request.scalars()

    score_entries = [f"{index + 1}. ID{item.user_id}: <b>{item.score}</b>" for index, item in enumerate(players)]
    score_entries_text = "\n".join(score_entries) \
        .replace(f"{message.from_user.id}", f"{message.from_user.id} (it's you!)")
    await message.answer(text_template.format(scores=score_entries_text), parse_mode="HTML")


def register_commands(dp: Dispatcher):
    dp.register_message_handler(cmd_start, commands="start")
    dp.register_message_handler(cmd_install, commands="install")
    dp.register_message_handler(get_card, commands="get")
    dp.register_message_handler(cmd_read_xlsx1, commands="rx1")
    dp.register_message_handler(cmd_read_xlsx2, commands="rx2")

    # dp.register_message_handler(cmd_top, commands="top")
